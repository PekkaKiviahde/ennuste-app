#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
import_jyda.py

MVP-import: Excel "Jyda-ajo" sheet -> Postgres
- Upsert litteras (project_id, code)
- Insert import_batches row (signature-based dedup)
- Insert actual_cost_lines as SNAPSHOT metrics (occurred_on = import date)

IMPORTANT:
Jyda-ajo is typically cumulative snapshot. Do NOT sum multiple imports.
Use views from migrations/0003_jyda_snapshot_views.sql for reporting the latest snapshot.

Usage example:
  python tools/scripts/import_jyda.py --file "excel/Tavo ja ennuste ....xlsm" --project-id <uuid> --db-url "postgresql://user:pass@127.0.0.1:5433/db" --imported-by Pekka

"""
from __future__ import annotations

import argparse
import csv
import getpass
import hashlib
import json
import os
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

try:
    import openpyxl
except ImportError as e:
    raise SystemExit("Missing dependency: openpyxl. Install with: pip install -r tools/scripts/requirements.txt") from e

try:
    import psycopg
except ImportError as e:
    raise SystemExit("Missing dependency: psycopg. Install with: pip install -r tools/scripts/requirements.txt") from e


CODE_RE = re.compile(r"^\d{4}$")

DEFAULT_METRICS = [
    "JYDA.ACTUAL_COST",                  # E
    "JYDA.COMMITTED_COST",               # D
    "JYDA.ACTUAL_COST_INCL_UNAPPROVED",  # F
    "JYDA.FORECAST_COST",                # G
    # "JYDA.TARGET_COST",                # C (optional; usually better into budget_lines later)
]

METRIC_TO_COLUMN_LETTER = {
    "JYDA.TARGET_COST": "C",
    "JYDA.COMMITTED_COST": "D",
    "JYDA.ACTUAL_COST": "E",
    "JYDA.ACTUAL_COST_INCL_UNAPPROVED": "F",
    "JYDA.FORECAST_COST": "G",
}

DEFAULT_JYDA_MAPPING = {
    "sheet_name": "Jyda-ajo",
    "code_column": "A",
    "name_column": "B",
    "metrics": METRIC_TO_COLUMN_LETTER,
    "csv_code_header": "Koodi",
    "csv_name_header": "Nimi",
    "csv_headers": {
        "JYDA.TARGET_COST": "Tavoitekustannus",
        "JYDA.COMMITTED_COST": "Sidottu kustannus",
        "JYDA.ACTUAL_COST": "Toteutunut kustannus",
        "JYDA.ACTUAL_COST_INCL_UNAPPROVED": "Toteutunut kustannus (sis. hyväksymätt.)",
        "JYDA.FORECAST_COST": "Ennustettu kustannus",
    },
}


@dataclass
class JydaRow:
    code: str
    name: str
    values: Dict[str, Optional[Decimal]]  # metric -> amount


def _sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _to_code(v) -> Optional[str]:
    """Normalize code to 4-digit string. Handles numeric cells that might lose leading zero."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        # Remove trailing .0 if user has it as "310.0"
        s = re.sub(r"\.0$", "", s)
        # If it's digits, keep as is; if numeric-like, parse to int and zero-pad
        if re.fullmatch(r"\d+", s):
            if len(s) < 4:
                s = s.zfill(4)
            return s
        return s  # might be 'L' etc. will be filtered out by regex later
    if isinstance(v, (int,)):
        return f"{v:04d}"
    if isinstance(v, float):
        if v.is_integer():
            return f"{int(v):04d}"
        return str(v)
    return str(v).strip()


def _to_decimal_money(v) -> Optional[Decimal]:
    """Parse Excel numeric or Finnish-formatted string into Decimal(2dp)."""
    if v is None:
        return None
    if isinstance(v, Decimal):
        return v.quantize(Decimal("0.01"))
    if isinstance(v, (int, float)):
        # Avoid float quirks by string conversion
        return Decimal(str(v)).quantize(Decimal("0.01"))
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        s = s.replace("\u00a0", "").replace(" ", "")  # nbsp and spaces
        s = s.replace("€", "")
        # Finnish: 1.234,56 or 1234,56
        if "," in s and "." in s:
            # assume '.' thousands, ',' decimal
            s = s.replace(".", "").replace(",", ".")
        elif "," in s and "." not in s:
            s = s.replace(",", ".")
        # else '.' decimal or plain digits
        try:
            return Decimal(s).quantize(Decimal("0.01"))
        except InvalidOperation:
            return None
    return None


def _excel_col_letter_to_index(letter: str) -> int:
    """A->1, B->2 ..."""
    letter = letter.strip().upper()
    n = 0
    for ch in letter:
        if not ("A" <= ch <= "Z"):
            raise ValueError(f"Invalid column letter: {letter}")
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n


def normalize_mapping(raw_mapping: Optional[Dict]) -> Dict[str, object]:
    if raw_mapping is None:
        raw_mapping = {}
    if "mapping" in raw_mapping and isinstance(raw_mapping["mapping"], dict):
        raw_mapping = raw_mapping["mapping"]
    mapping = {
        "sheet_name": DEFAULT_JYDA_MAPPING["sheet_name"],
        "code_column": DEFAULT_JYDA_MAPPING["code_column"],
        "name_column": DEFAULT_JYDA_MAPPING["name_column"],
        "metrics": dict(DEFAULT_JYDA_MAPPING["metrics"]),
        "csv_code_header": DEFAULT_JYDA_MAPPING["csv_code_header"],
        "csv_name_header": DEFAULT_JYDA_MAPPING["csv_name_header"],
        "csv_headers": dict(DEFAULT_JYDA_MAPPING["csv_headers"]),
    }
    if isinstance(raw_mapping.get("sheet_name"), str) and raw_mapping["sheet_name"].strip():
        mapping["sheet_name"] = raw_mapping["sheet_name"].strip()
    if isinstance(raw_mapping.get("code_column"), str) and raw_mapping["code_column"].strip():
        mapping["code_column"] = raw_mapping["code_column"].strip()
    if isinstance(raw_mapping.get("name_column"), str) and raw_mapping["name_column"].strip():
        mapping["name_column"] = raw_mapping["name_column"].strip()
    if isinstance(raw_mapping.get("metrics"), dict):
        for key, value in raw_mapping["metrics"].items():
            if isinstance(value, str) and value.strip():
                mapping["metrics"][key] = value.strip().upper()
    if isinstance(raw_mapping.get("csv_code_header"), str) and raw_mapping["csv_code_header"].strip():
        mapping["csv_code_header"] = raw_mapping["csv_code_header"].strip()
    if isinstance(raw_mapping.get("csv_name_header"), str) and raw_mapping["csv_name_header"].strip():
        mapping["csv_name_header"] = raw_mapping["csv_name_header"].strip()
    if isinstance(raw_mapping.get("csv_headers"), dict):
        for key, value in raw_mapping["csv_headers"].items():
            if isinstance(value, str) and value.strip():
                mapping["csv_headers"][key] = value.strip()
    return mapping


def load_mapping_from_db(conn, project_id: str) -> Optional[Dict[str, object]]:
    with conn.cursor() as cur:
        cur.execute(
            "SELECT mapping FROM import_mappings WHERE project_id=%s AND import_type='JYDA'",
            (project_id,),
        )
        row = cur.fetchone()
        if not row:
            return None
        if isinstance(row[0], dict):
            return row[0]
        return None


def read_jyda_from_excel(
    path: Path,
    sheet_name: str,
    metrics: List[str],
    metric_to_column: Dict[str, str],
    code_column: str,
    name_column: str,
) -> List[JydaRow]:
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Sheets: {wb.sheetnames}")
    ws = wb[sheet_name]

    # Validate header row (optional but helpful)
    header = [ws.cell(row=1, column=c).value for c in range(1, 15)]
    if not header or header[0] != "Koodi":
        # not fatal; but warn
        print(f"WARNING: Header row doesn't look like expected Jyda-ajo. A1..N1={header}", file=sys.stderr)

    rows: List[JydaRow] = []
    max_row = ws.max_row or 0

    code_col_idx = _excel_col_letter_to_index(code_column)
    name_col_idx = _excel_col_letter_to_index(name_column)

    # Precompute metric->column index
    metric_cols: Dict[str, int] = {}
    for m in metrics:
        col_letter = metric_to_column.get(m)
        if not col_letter:
            raise ValueError(f"Unknown metric '{m}'. Known: {sorted(METRIC_TO_COLUMN_LETTER.keys())}")
        metric_cols[m] = _excel_col_letter_to_index(col_letter)

    for r in range(2, max_row + 1):
        raw_code = ws.cell(row=r, column=code_col_idx).value
        raw_name = ws.cell(row=r, column=name_col_idx).value

        code = _to_code(raw_code)
        if code is None:
            continue
        if not CODE_RE.match(code):
            # skip totals like 'L', 'LSOS', blanks, etc.
            continue

        name = str(raw_name).strip() if raw_name is not None else ""

        values: Dict[str, Optional[Decimal]] = {}
        for metric, cidx in metric_cols.items():
            v = ws.cell(row=r, column=cidx).value
            values[metric] = _to_decimal_money(v)

        rows.append(JydaRow(code=code, name=name, values=values))

    wb.close()
    return rows


def read_jyda_from_csv(
    path: Path,
    metrics: List[str],
    metric_to_header: Dict[str, str],
    code_header: str,
    name_header: str,
) -> List[JydaRow]:
    """
    CSV input option. Expects columns matching Jyda-ajo headers:
    Koodi, Nimi, Tavoitekustannus, Sidottu kustannus, Toteutunut kustannus, ...
    """
    rows: List[JydaRow] = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter=",")
        if reader.fieldnames is None:
            raise ValueError("CSV has no header row.")
        required_headers = [code_header, name_header] + [metric_to_header[m] for m in metrics if m in metric_to_header]
        missing_headers = [h for h in required_headers if h not in reader.fieldnames]
        if missing_headers:
            raise ValueError(f"CSV missing required columns: {missing_headers}")
        for rec in reader:
            code = _to_code(rec.get(code_header))
            if code is None or not CODE_RE.match(code):
                continue
            name = (rec.get(name_header) or "").strip()
            values: Dict[str, Optional[Decimal]] = {}
            for m in metrics:
                hdr = metric_to_header.get(m)
                if not hdr:
                    raise ValueError(f"Unknown metric '{m}' for CSV")
                values[m] = _to_decimal_money(rec.get(hdr))
            rows.append(JydaRow(code=code, name=name, values=values))
    return rows


def connect(db_url: str):
    return psycopg.connect(db_url)


def ensure_project_exists(cur, project_id: str) -> None:
    cur.execute("SELECT 1 FROM projects WHERE project_id = %s", (project_id,))
    if cur.fetchone() is None:
        raise RuntimeError(f"Project not found in DB: {project_id}. Create it first in projects table.")


def import_batch_exists(cur, project_id: str, signature: str) -> bool:
    cur.execute(
        "SELECT 1 FROM import_batches WHERE project_id = %s AND signature = %s LIMIT 1",
        (project_id, signature),
    )
    return cur.fetchone() is not None


def insert_import_batch(cur, project_id: str, imported_by: str, signature: str, notes: str) -> str:
    cur.execute(
        """
        INSERT INTO import_batches (project_id, source_system, imported_by, signature, notes)
        VALUES (%s, %s, %s, %s, %s)
        RETURNING import_batch_id
        """,
        (project_id, "JYDA", imported_by, signature, notes),
    )
    return cur.fetchone()[0]


def upsert_littera(cur, project_id: str, code: str, title: str) -> str:
    group_code = int(code[0]) if code and code[0].isdigit() else None
    cur.execute(
        """
        INSERT INTO litteras (project_id, code, title, group_code, is_active)
        VALUES (%s, %s, %s, %s, true)
        ON CONFLICT (project_id, code)
        DO UPDATE SET
          title = EXCLUDED.title,
          group_code = COALESCE(EXCLUDED.group_code, litteras.group_code),
          is_active = true
        RETURNING littera_id
        """,
        (project_id, code, title, group_code),
    )
    return cur.fetchone()[0]


def insert_actual_line(cur, project_id: str, littera_id: str, amount: Decimal, occurred_on: date, import_batch_id: str, external_ref: str) -> None:
    cur.execute(
        """
        INSERT INTO actual_cost_lines (
          project_id, work_littera_id, cost_type, amount, occurred_on, source, import_batch_id, external_ref
        )
        VALUES (%s, %s, 'OTHER', %s, %s, 'JYDA', %s, %s)
        """,
        (project_id, littera_id, amount, occurred_on, import_batch_id, external_ref),
    )


def parse_args(argv: List[str]) -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Import Jyda-ajo from Excel/CSV into Postgres (MVP).")
    p.add_argument("--file", required=True, help="Path to .xlsm/.xlsx or .csv")
    p.add_argument("--sheet", default="Jyda-ajo", help="Excel sheet name (default: Jyda-ajo)")
    p.add_argument("--project-id", required=True, help="UUID of projects.project_id")
    p.add_argument("--db-url", default=os.environ.get("DATABASE_URL", ""), help="Postgres URL. Or set env DATABASE_URL.")
    p.add_argument("--imported-by", default=os.environ.get("IMPORTED_BY", getpass.getuser()), help="Name of importer (audit)")
    p.add_argument("--occurred-on", default=os.environ.get("OCCURRED_ON", ""), help="Snapshot date YYYY-MM-DD (default: today)")
    p.add_argument("--metrics", nargs="*", default=DEFAULT_METRICS, help=f"Metrics to import. Default: {DEFAULT_METRICS}")
    p.add_argument("--mapping-file", help="Path to JSON mapping file (optional).")
    p.add_argument("--mapping-json", help="Inline JSON mapping (optional).")
    p.add_argument(
        "--mapping-source",
        choices=["auto", "db", "none"],
        default="auto",
        help="Mapping source: auto (file/json -> db -> default), db, or none (defaults only).",
    )
    p.add_argument("--include-zeros", action="store_true", help="Insert rows even if amount=0.00 (default: skip zeros).")
    p.add_argument("--dry-run", action="store_true", help="Run validations and print summary but ROLLBACK at end.")
    return p.parse_args(argv)


def main(argv: List[str]) -> int:
    args = parse_args(argv)

    file_path = Path(args.file).expanduser().resolve()
    if not file_path.exists():
        print(f"ERROR: file not found: {file_path}", file=sys.stderr)
        return 2

    if not args.db_url:
        print("ERROR: Missing --db-url or env DATABASE_URL", file=sys.stderr)
        return 2

    if args.occurred_on:
        occurred_on = datetime.strptime(args.occurred_on, "%Y-%m-%d").date()
    else:
        occurred_on = date.today()

    conn = connect(args.db_url)
    conn.autocommit = False
    try:
        raw_mapping: Optional[Dict[str, object]] = None
        if args.mapping_source in ("auto", "db"):
            raw_mapping = load_mapping_from_db(conn, args.project_id)
        if args.mapping_file:
            try:
                raw_mapping = json.loads(Path(args.mapping_file).read_text(encoding="utf-8"))
            except Exception as e:
                print(f"ERROR reading mapping file: {e}", file=sys.stderr)
                return 2
        if args.mapping_json:
            try:
                raw_mapping = json.loads(args.mapping_json)
            except Exception as e:
                print(f"ERROR parsing mapping JSON: {e}", file=sys.stderr)
                return 2
        mapping = normalize_mapping(raw_mapping)

        effective_sheet = args.sheet if args.sheet != "Jyda-ajo" else mapping["sheet_name"]
        metric_to_column = mapping["metrics"]
        code_column = mapping["code_column"]
        name_column = mapping["name_column"]
        csv_headers = mapping["csv_headers"]
        csv_code_header = mapping["csv_code_header"]
        csv_name_header = mapping["csv_name_header"]

        # Compute signature: sha256(file bytes) + sheet name + metrics list + mapping fingerprint
        file_hash = _sha256_file(file_path)
        mapping_fingerprint = json.dumps(mapping, sort_keys=True)
        signature = hashlib.sha256(
            (file_hash + "|" + effective_sheet + "|" + ",".join(args.metrics) + "|" + mapping_fingerprint).encode("utf-8")
        ).hexdigest()

        # Read rows
        try:
            if file_path.suffix.lower() == ".csv":
                jyda_rows = read_jyda_from_csv(
                    file_path,
                    metrics=args.metrics,
                    metric_to_header=csv_headers,
                    code_header=csv_code_header,
                    name_header=csv_name_header,
                )
            else:
                jyda_rows = read_jyda_from_excel(
                    file_path,
                    sheet_name=str(effective_sheet),
                    metrics=args.metrics,
                    metric_to_column=metric_to_column,
                    code_column=str(code_column),
                    name_column=str(name_column),
                )
        except Exception as e:
            print(f"ERROR reading source: {e}", file=sys.stderr)
            return 2
    finally:
        conn.close()

    if not jyda_rows:
        print("No rows found to import (after filtering).", file=sys.stderr)
        return 1

    # Connect DB
    conn = connect(args.db_url)
    conn.autocommit = False
    try:
        with conn.cursor() as cur:
            ensure_project_exists(cur, args.project_id)

            if import_batch_exists(cur, args.project_id, signature):
                print("Import already done (same signature found). Aborting to avoid duplicates.")
                conn.rollback()
                return 0

            notes = f"JYDA snapshot import from {file_path.name} (sheet={effective_sheet}, metrics={args.metrics})"
            import_batch_id = insert_import_batch(cur, args.project_id, args.imported_by, signature, notes)

            littera_count = 0
            metric_inserts = 0
            skipped_zero = 0

            for row in jyda_rows:
                littera_id = upsert_littera(cur, args.project_id, row.code, row.name)
                littera_count += 1

                for metric, amount in row.values.items():
                    if amount is None:
                        continue
                    if (not args.include_zeros) and amount == Decimal("0.00"):
                        skipped_zero += 1
                        continue
                    insert_actual_line(cur, args.project_id, littera_id, amount, occurred_on, import_batch_id, metric)
                    metric_inserts += 1

            # Summary
            print("=== Import summary ===")
            print(f"Project:        {args.project_id}")
            print(f"Source file:    {file_path}")
            print(f"Sheet:          {args.sheet}")
            print(f"Occurred_on:    {occurred_on.isoformat()} (snapshot date)")
            print(f"Metrics:        {args.metrics}")
            print(f"Rows (littera): {littera_count}")
            print(f"Inserted lines: {metric_inserts}")
            print(f"Skipped zeros:  {skipped_zero}")
            print(f"Import batch:   {import_batch_id}")
            print(f"Signature:      {signature[:12]}...")

            if args.dry_run:
                print("DRY-RUN enabled -> ROLLBACK")
                conn.rollback()
            else:
                conn.commit()
                print("COMMIT OK")

        return 0
    except Exception as e:
        conn.rollback()
        print(f"ERROR importing: {e}", file=sys.stderr)
        return 3
    finally:
        conn.close()


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
