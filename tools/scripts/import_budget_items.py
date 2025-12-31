#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
import_budget_items.py

Imports detailed target estimate line items (777-ish rows) into a table "budget_items".

Excel layout (your confirmed headers):
A Litterakoodi, B Litteraselite, C Koodi, D Selite, E Määrä, F Yksikkö,
G Työ €/h, H Työ €/yks., I Työ €, J Aine €/yks., K Aine €, L Alih €/yks., M Alih €,
N Vmiehet €/yks., O Vmiehet €, P Muu €, Q Summa.

Purpose:
- Learning/audit: determine whether a detailed item existed in target estimate at calculation stage.
- Link items to an existing TARGET_ESTIMATE import_batch_id (defaults to latest for the project).

Usage:
  python tools/scripts/import_budget_items.py --file "excel/Tavoitearvio ....xlsx" --project-id <uuid> --dry-run
  python tools/scripts/import_budget_items.py --file "excel/Tavoitearvio ....xlsx" --project-id <uuid>

Before running:
- You must create table budget_items (migration 0004). If you don’t have it yet, tell me and I’ll paste it too.
"""

from __future__ import annotations

import argparse
import getpass
import os
import re
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional, List

try:
    import openpyxl
except ImportError as e:
    raise SystemExit("Missing dependency: openpyxl. Install with: pip install -r tools/scripts/requirements.txt") from e

try:
    import psycopg
except ImportError as e:
    raise SystemExit("Missing dependency: psycopg. Install with: pip install -r tools/scripts/requirements.txt") from e


CODE4_RE = re.compile(r"^\d{4}$")

# Fixed column mapping based on your header row (A..Q)
COL = {
    "code4": 1, "name4": 2, "item_code": 3, "item_desc": 4,
    "qty": 5, "unit": 6,
    "labor_unit_price": 8,      # H
    "labor_eur": 9,             # I
    "material_unit_price": 10,  # J
    "material_eur": 11,         # K
    "subcontract_unit_price": 12,  # L
    "subcontract_eur": 13,         # M
    "rental_unit_price": 14,       # N
    "rental_eur": 15,              # O
    "other_eur": 16,               # P
    "total_eur": 17,               # Q
}


def _to_code(v) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        s = re.sub(r"\.0$", "", s)
        if re.fullmatch(r"\d+", s):
            if len(s) < 4:
                s = s.zfill(4)
            return s
        return s
    if isinstance(v, int):
        return f"{v:04d}"
    if isinstance(v, float) and v.is_integer():
        return f"{int(v):04d}"
    return str(v).strip()


def _to_dec(v) -> Optional[Decimal]:
    if v is None:
        return None
    if isinstance(v, Decimal):
        return v
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        s = s.replace("\u00a0", "").replace(" ", "").replace("€", "")
        # Finnish formatting: 1.234,56
        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")
        elif "," in s and "." not in s:
            s = s.replace(",", ".")
        try:
            return Decimal(s)
        except InvalidOperation:
            return None
    return None


def connect(db_url: str):
    return psycopg.connect(db_url)


def ensure_project(cur, project_id: str):
    cur.execute("SELECT 1 FROM projects WHERE project_id=%s", (project_id,))
    if cur.fetchone() is None:
        raise RuntimeError(f"Project not found: {project_id}")


def find_latest_target_estimate_batch(cur, project_id: str) -> str:
    cur.execute(
        """
        SELECT import_batch_id
        FROM import_batches
        WHERE project_id=%s AND source_system='TARGET_ESTIMATE'
        ORDER BY imported_at DESC
        LIMIT 1
        """,
        (project_id,),
    )
    row = cur.fetchone()
    if not row:
        raise RuntimeError("No TARGET_ESTIMATE import batch found for project. Run budget import first.")
    return row[0]


def get_littera_id(cur, project_id: str, code4: str, title: str) -> str:
    group_code = int(code4[0])
    cur.execute(
        """
        INSERT INTO litteras (project_id, code, title, group_code, is_active)
        VALUES (%s, %s, %s, %s, true)
        ON CONFLICT (project_id, code)
        DO UPDATE SET
          title = CASE WHEN EXCLUDED.title IS NOT NULL AND EXCLUDED.title <> '' THEN EXCLUDED.title ELSE litteras.title END,
          group_code = COALESCE(EXCLUDED.group_code, litteras.group_code),
          is_active = true
        RETURNING littera_id
        """,
        (project_id, code4, title, group_code),
    )
    return cur.fetchone()[0]


def insert_item(
    cur,
    project_id: str,
    batch_id: str,
    littera_id: str,
    row_no: int,
    item_code: str,
    item_desc: str,
    qty,
    unit: str,
    lup, mup, sup, rup,
    le, me, se, reu, oe, te,
    created_by: str,
):
    cur.execute(
        """
        INSERT INTO budget_items (
          project_id, import_batch_id, littera_id, item_code, item_desc, row_no,
          qty, unit,
          labor_unit_price, material_unit_price, subcontract_unit_price, rental_unit_price,
          labor_eur, material_eur, subcontract_eur, rental_eur, other_eur, total_eur,
          created_by
        ) VALUES (
          %s,%s,%s,%s,%s,%s,
          %s,%s,
          %s,%s,%s,%s,
          %s,%s,%s,%s,%s,%s,
          %s
        )
        ON CONFLICT (import_batch_id, row_no) DO NOTHING
        """,
        (
            project_id, batch_id, littera_id, item_code, item_desc, row_no,
            qty, unit,
            lup, mup, sup, rup,
            le, me, se, reu, oe, te,
            created_by
        ),
    )


def parse_args(argv: List[str]) -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Import detailed target estimate items into budget_items.")
    p.add_argument("--file", required=True, help="Path to target estimate .xlsx/.xlsm")
    p.add_argument("--sheet", default="", help="Sheet name (default: first sheet)")
    p.add_argument("--project-id", required=True, help="projects.project_id UUID")
    p.add_argument("--db-url", default=os.environ.get("DATABASE_URL", ""), help="Postgres URL or env DATABASE_URL")
    p.add_argument("--imported-by", default=os.environ.get("IMPORTED_BY", getpass.getuser()))
    p.add_argument("--import-batch-id", default="", help="Existing TARGET_ESTIMATE import_batch_id to link items to.")
    p.add_argument("--dry-run", action="store_true")
    return p.parse_args(argv)


def main(argv: List[str]) -> int:
    args = parse_args(argv)
    fp = Path(args.file).expanduser().resolve()
    if not fp.exists():
        print(f"ERROR: file not found: {fp}", file=sys.stderr)
        return 2
    if not args.db_url:
        print("ERROR: Missing DATABASE_URL (set it like before).", file=sys.stderr)
        return 2

    wb = openpyxl.load_workbook(str(fp), data_only=True, read_only=True)
    sheet = args.sheet or wb.sheetnames[0]
    if sheet not in wb.sheetnames:
        print(f"ERROR: sheet not found: {sheet}. Sheets: {wb.sheetnames}", file=sys.stderr)
        return 2
    ws = wb[sheet]

    conn = connect(args.db_url)
    conn.autocommit = False
    try:
        with conn.cursor() as cur:
            ensure_project(cur, args.project_id)
            batch_id = args.import_batch_id or find_latest_target_estimate_batch(cur, args.project_id)

            inserted = 0
            littera_upserts = 0

            max_row = ws.max_row or 0
            for r in range(2, max_row + 1):
                code4 = _to_code(ws.cell(row=r, column=COL["code4"]).value)
                if not code4 or not CODE4_RE.match(code4):
                    continue

                name4 = ws.cell(row=r, column=COL["name4"]).value
                name4 = str(name4).strip() if name4 is not None else ""
                littera_id = get_littera_id(cur, args.project_id, code4, name4)
                littera_upserts += 1

                item_code = ws.cell(row=r, column=COL["item_code"]).value
                item_code = str(item_code).strip() if item_code is not None else ""
                item_desc = ws.cell(row=r, column=COL["item_desc"]).value
                item_desc = str(item_desc).strip() if item_desc is not None else ""

                qty = _to_dec(ws.cell(row=r, column=COL["qty"]).value)
                unit = ws.cell(row=r, column=COL["unit"]).value
                unit = str(unit).strip() if unit is not None else ""

                lup = _to_dec(ws.cell(row=r, column=COL["labor_unit_price"]).value)
                mup = _to_dec(ws.cell(row=r, column=COL["material_unit_price"]).value)
                sup = _to_dec(ws.cell(row=r, column=COL["subcontract_unit_price"]).value)
                rup = _to_dec(ws.cell(row=r, column=COL["rental_unit_price"]).value)

                le = _to_dec(ws.cell(row=r, column=COL["labor_eur"]).value)
                me = _to_dec(ws.cell(row=r, column=COL["material_eur"]).value)
                se = _to_dec(ws.cell(row=r, column=COL["subcontract_eur"]).value)
                reu = _to_dec(ws.cell(row=r, column=COL["rental_eur"]).value)
                oe = _to_dec(ws.cell(row=r, column=COL["other_eur"]).value)
                te = _to_dec(ws.cell(row=r, column=COL["total_eur"]).value)

                # Skip fully empty money rows
                if all(x is None or x == 0 for x in [le, me, se, reu, oe, te]):
                    continue

                insert_item(
                    cur, args.project_id, batch_id, littera_id, r, item_code, item_desc,
                    qty, unit, lup, mup, sup, rup, le, me, se, reu, oe, te, args.imported_by
                )
                inserted += 1

            print("=== Budget items import summary ===")
            print(f"Project:        {args.project_id}")
            print(f"Source file:    {fp}")
            print(f"Sheet:          {sheet}")
            print(f"Linked batch:   {batch_id} (TARGET_ESTIMATE)")
            print(f"Inserted items: {inserted}")
            print(f"Littera upserts:{littera_upserts}")

            if args.dry_run:
                print("DRY-RUN enabled -> ROLLBACK")
                conn.rollback()
            else:
                conn.commit()
                print("COMMIT OK")

        return 0
    except Exception as e:
        conn.rollback()
        print(f"ERROR importing budget items: {e}", file=sys.stderr)
        return 3
    finally:
        conn.close()
        wb.close()


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
